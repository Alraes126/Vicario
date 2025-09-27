# controllers/bet_controller.py
# Este archivo define el controlador para la gestión de apuestas.
# Un controlador actúa como intermediario entre la Vista (lo que el usuario ve)
# y el Modelo (la lógica de datos y negocio).

# --- Importación de Bibliotecas ---
from fpdf import FPDF # Importamos FPDF para generar documentos PDF.
import openpyxl       # Importamos openpyxl para trabajar con archivos Excel (.xlsx).
from tkinter import messagebox # Para mostrar mensajes emergentes al usuario.

# --- Definición de la Clase BetController ---
# Esta clase sigue el principio de Responsabilidad Única (SRP)
# al encargarse exclusivamente de la lógica relacionada con las apuestas.
class BetController:
    # El constructor (__init__) inicializa el controlador con las dependencias necesarias.
    # Esto es un ejemplo de Inyección de Dependencias.
    def __init__(self, view, bet_model, user_model, game_model):
        self.view = view             # La Vista asociada a este controlador (ej. BetsWindow).
        self.bet_model = bet_model   # El Modelo de Apuestas para interactuar con los datos de apuestas.
        self.user_model = user_model # El Modelo de Usuario para obtener información del usuario.
        self.game_model = game_model # El Modelo de Juego para obtener detalles de los juegos.
        self.current_user = None     # Almacena los datos del usuario actualmente logueado.

    # Método para establecer el usuario actual en el controlador.
    # Se llama cuando un usuario inicia sesión.
    def set_current_user(self, user_data):
        self.current_user = user_data
        self.load_user_bets() # Carga las apuestas del usuario una vez que se establece.

    # Método para cargar y mostrar las apuestas del usuario, con opción de filtrar por fecha.
    def load_user_bets(self, start_date=None, end_date=None):
        if self.current_user: # Verificamos que haya un usuario logueado.
            # Obtenemos las apuestas del modelo, aplicando filtros de fecha si se proporcionan.
            bets = self.bet_model.get_bets_by_user(self.current_user['idcedula'], start_date, end_date)
            
            # --- Mejora de Datos (Enriquecimiento) ---
            # Para cada apuesta, obtenemos el nombre del juego asociado para mostrarlo en la vista.
            enhanced_bets = []
            for bet in bets:
                game = self.game_model.get_game_by_id(bet['idjuego'])
                bet['nombre_juego'] = game['nombre'] if game else 'Desconocido' # Asignamos el nombre del juego.
                enhanced_bets.append(bet)
            
            # Le decimos a la Vista que muestre las apuestas mejoradas.
            self.view.display_bets(enhanced_bets)

    # Método para exportar la lista de apuestas a un archivo PDF.
    # Utiliza la librería FPDF para crear el documento.
    def export_bets_to_pdf(self, bets, filename="bets_report.pdf"):
        if not bets: # Si no hay apuestas, mostramos un mensaje y salimos.
            messagebox.showinfo("Exportar PDF", "No hay apuestas para exportar.")
            return

        pdf = FPDF()         # Creamos una nueva instancia de FPDF.
        pdf.add_page()       # Añadimos una página al documento.
        pdf.set_font("Arial", size=12) # Establecemos la fuente y el tamaño.

        # Añadimos un título al reporte.
        pdf.cell(200, 10, txt="Reporte de Apuestas", ln=True, align="C")
        pdf.ln(10) # Añadimos un salto de línea.

        # --- Encabezados de la Tabla PDF ---
        pdf.set_font("Arial", size=10, style='B') # Fuente en negrita para los encabezados.
        headers = ["ID", "Juego", "Monto", "Resultado", "Ganancia", "Fecha"]
        col_widths = [15, 30, 25, 25, 25, 40] # Anchos de columna definidos manualmente.
        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], 7, header, border=1, align="C") # Creamos una celda para cada encabezado.
        pdf.ln() # Salto de línea después de los encabezados.

        # --- Datos de la Tabla PDF ---
        pdf.set_font("Arial", size=8) # Fuente normal para los datos.
        for bet in bets: # Iteramos sobre cada apuesta para añadirla a la tabla.
            pdf.cell(col_widths[0], 7, str(bet['idapuesta']), border=1)
            pdf.cell(col_widths[1], 7, bet['nombre_juego'], border=1)
            pdf.cell(col_widths[2], 7, f"${bet['monto']:.2f}", border=1)
            pdf.cell(col_widths[3], 7, str(bet['resultado']), border=1)
            pdf.cell(col_widths[4], 7, f"${bet['ganancia']:.2f}", border=1)
            pdf.cell(col_widths[5], 7, str(bet['fecha_apuesta']), border=1)
            pdf.ln() # Salto de línea después de cada fila de datos.

        try:
            pdf.output(filename) # Guardamos el documento PDF en el archivo especificado.
            messagebox.showinfo("Exportar PDF", f"Reporte de apuestas exportado a {filename}")
        except Exception as e:
            messagebox.showerror("Error de Exportación", f"No se pudo exportar a PDF: {e}")

    # Método para exportar la lista de apuestas a un archivo Excel (.xlsx).
    # Utiliza la librería openpyxl para crear el libro de trabajo.
    def export_bets_to_excel(self, bets, filename="bets_report.xlsx"):
        if not bets: # Si no hay apuestas, mostramos un mensaje y salimos.
            messagebox.showinfo("Exportar Excel", "No hay apuestas para exportar.")
            return

        workbook = openpyxl.Workbook() # Creamos un nuevo libro de trabajo de Excel.
        sheet = workbook.active        # Obtenemos la hoja activa (la primera por defecto).
        sheet.title = "Reporte de Apuestas" # Establecemos el título de la hoja.

        # --- Encabezados de la Tabla Excel ---
        headers = ["ID", "Juego", "Monto", "Resultado", "Ganancia", "Fecha"]
        sheet.append(headers) # Añadimos los encabezados como la primera fila.

        # --- Datos de la Tabla Excel ---
        for bet in bets: # Iteramos sobre cada apuesta para añadirla a la hoja.
            sheet.append([
                bet['idapuesta'],
                bet['nombre_juego'],
                bet['monto'],
                bet['resultado'],
                bet['ganancia'],
                bet['fecha_apuesta']
            ])
        try:
            workbook.save(filename) # Guardamos el libro de trabajo de Excel en el archivo especificado.
            messagebox.showinfo("Exportar Excel", f"Reporte de apuestas exportado a {filename}")
        except Exception as e:
            messagebox.showerror("Error de Exportación", f"No se pudo exportar a Excel: {e}")