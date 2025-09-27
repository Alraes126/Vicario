# controllers/transaction_controller.py
# Este archivo define el controlador para la gestión de transacciones (depósitos).
# Un controlador gestiona la interacción del usuario con la vista de transacciones
# y coordina con los modelos de usuario y transacciones para procesar operaciones monetarias.

# --- Importación de Bibliotecas ---
from tkinter import messagebox # Para mostrar mensajes emergentes al usuario.
from decimal import Decimal # Importamos 'Decimal' para manejar cálculos monetarios con precisión.
from fpdf import FPDF       # Importamos FPDF para generar documentos PDF.
import openpyxl             # Importamos openpyxl para trabajar con archivos Excel (.xlsx).

# --- Definición de la Clase TransactionController ---
# Esta clase es un ejemplo del patrón de diseño MVC (Modelo-Vista-Controlador).
# Su responsabilidad es manejar la lógica de las transacciones, especialmente los depósitos.
class TransactionController:
    # El constructor (__init__) inicializa el controlador con las dependencias necesarias.
    # Esto es un ejemplo de Inyección de Dependencias.
    def __init__(self, view, transaction_model, user_model):
        self.view = view                     # La Vista asociada a este controlador (TransactionsWindow).
        self.transaction_model = transaction_model # El Modelo de Transacciones para interactuar con los datos de transacciones.
        self.user_model = user_model         # El Modelo de Usuario para interactuar con los datos del usuario (saldo).
        self.current_user = None             # Almacena los datos del usuario actualmente logueado.
        self.dashboard_controller = None     # Referencia al controlador del Dashboard para actualizar el saldo.

    # Método para establecer el usuario actual en el controlador.
    # Se llama cuando un usuario inicia sesión o cuando los datos del usuario se actualizan.
    def set_current_user(self, user_data):
        self.current_user = user_data       # Actualizamos el usuario actual en este controlador.
        self.load_user_transactions() # Carga las transacciones del usuario una vez que se establece.

    # Método para cargar y mostrar las transacciones del usuario, con opción de filtrar por fecha.
    def load_user_transactions(self, start_date=None, end_date=None):
        if self.current_user: # Verificamos que haya un usuario logueado.
            # Obtenemos las transacciones del modelo, aplicando filtros de fecha si se proporcionan.
            transactions = self.transaction_model.get_transactions_by_user(self.current_user['idcedula'], start_date, end_date)
            # Le decimos a la Vista que muestre las transacciones.
            self.view.display_transactions(transactions)

    # Método para procesar una solicitud de depósito.
    # Recibe el monto del depósito como cadena de texto y el método de pago.
    def request_deposit(self, amount_str, payment_method):
        if not self.current_user: # Verificamos que haya un usuario logueado.
            messagebox.showerror("Error", "No hay usuario logueado para realizar un depósito.")
            return

        # --- Validación del Monto del Depósito ---
        try:
            amount = Decimal(amount_str) # Convertimos el monto a tipo Decimal para cálculos precisos.
            if amount <= 0: # El monto debe ser positivo.
                messagebox.showerror("Error", "El monto del depósito debe ser positivo.")
                return
        except Exception: # Capturamos errores si el monto no es un número válido.
            messagebox.showerror("Error", "Monto inválido. Introduce un número válido.")
            return

        # --- Creación del Registro de Transacción ---
        # Preparamos los datos para crear una nueva transacción.
        transaction_data = {
            'idcedula': self.current_user['idcedula'],
            'tipo': 'deposito',
            'metododepago': payment_method,
            'monto_transaccion': amount,
            'estado': 'completado' # Asumimos que los depósitos se completan instantáneamente.
        }
        # Llamamos al modelo de transacciones para registrar la transacción en la base de datos.
        transaction_success = self.transaction_model.create_transaction(transaction_data)

        if transaction_success: # Si la transacción se registró exitosamente...
            # --- Actualización del Saldo del Usuario ---
            # Calculamos el nuevo saldo del usuario.
            new_balance = self.current_user['saldo'] + amount
            # Actualizamos el saldo en la base de datos a través del modelo de usuario.
            user_balance_updated = self.user_model.update_user_balance(self.current_user['idcedula'], new_balance)

            if user_balance_updated: # Si el saldo del usuario se actualizó correctamente...
                self.current_user['saldo'] = new_balance # Actualizamos el saldo en los datos locales del usuario.
                messagebox.showinfo("Éxito", f"Depósito de ${amount:.2f} realizado con éxito. Nuevo saldo: ${new_balance:.2f}")
                self.view.load_transactions() # Le decimos a la Vista que refresque la lista de transacciones.
                if self.dashboard_controller: # Si el controlador del dashboard está disponible, lo actualizamos.
                    self.dashboard_controller.refresh_user_data()
            else: # Si el saldo no se pudo actualizar.
                messagebox.showerror("Error", "Depósito registrado, pero no se pudo actualizar el saldo del usuario.")
        else: # Si la transacción no se pudo registrar.
            messagebox.showerror("Error", "No se pudo registrar el depósito.")

    # Método para exportar la lista de transacciones a un archivo PDF.
    # Utiliza la librería FPDF para crear el documento.
    def export_transactions_to_pdf(self, transactions, filename="transactions_report.pdf"):
        if not transactions: # Si no hay transacciones, mostramos un mensaje y salimos.
            messagebox.showinfo("Exportar PDF", "No hay transacciones para exportar.")
            return

        pdf = FPDF()         # Creamos una nueva instancia de FPDF.
        pdf.add_page()       # Añadimos una página al documento.
        pdf.set_font("Arial", size=12) # Establecemos la fuente y el tamaño.

        # Añadimos un título al reporte.
        pdf.cell(200, 10, txt="Reporte de Transacciones", ln=True, align="C")
        pdf.ln(10) # Añadimos un salto de línea.

        # --- Encabezados de la Tabla PDF ---
        pdf.set_font("Arial", size=10, style='B') # Fuente en negrita para los encabezados.
        headers = ["ID", "Tipo", "Método Pago", "Monto", "Fecha", "Estado"]
        col_widths = [15, 25, 35, 25, 40, 25] # Anchos de columna definidos manualmente.
        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], 7, header, border=1, align="C") # Creamos una celda para cada encabezado.
        pdf.ln() # Salto de línea después de los encabezados.

        # --- Datos de la Tabla PDF ---
        pdf.set_font("Arial", size=8) # Fuente normal para los datos.
        for trans in transactions: # Iteramos sobre cada transacción para añadirla a la tabla.
            pdf.cell(col_widths[0], 7, str(trans['idtransaccion']), border=1)
            pdf.cell(col_widths[1], 7, trans['tipo'], border=1)
            pdf.cell(col_widths[2], 7, trans['metododepago'], border=1)
            pdf.cell(col_widths[3], 7, f"${trans['monto_transaccion']:.2f}", border=1)
            pdf.cell(col_widths[4], 7, str(trans['fecha_transaccion']), border=1)
            pdf.cell(col_widths[5], 7, trans['estado'], border=1)
            pdf.ln() # Salto de línea después de cada fila de datos.

        try:
            pdf.output(filename) # Guardamos el documento PDF en el archivo especificado.
            messagebox.showinfo("Exportar PDF", f"Reporte de transacciones exportado a {filename}")
        except Exception as e:
            messagebox.showerror("Error de Exportación", f"No se pudo exportar a PDF: {e}")

    # Método para exportar la lista de transacciones a un archivo Excel (.xlsx).
    # Utiliza la librería openpyxl para crear el libro de trabajo.
    def export_transactions_to_excel(self, transactions, filename="transactions_report.xlsx"):
        if not transactions: # Si no hay transacciones, mostramos un mensaje y salimos.
            messagebox.showinfo("Exportar Excel", "No hay transacciones para exportar.")
            return

        workbook = openpyxl.Workbook() # Creamos un nuevo libro de trabajo de Excel.
        sheet = workbook.active        # Obtenemos la hoja activa (la primera por defecto).
        sheet.title = "Reporte de Transacciones" # Establecemos el título de la hoja.

        # --- Encabezados de la Tabla Excel ---
        headers = ["ID", "Tipo", "Método Pago", "Monto", "Fecha", "Estado"]
        sheet.append(headers) # Añadimos los encabezados como la primera fila.

        # --- Datos de la Tabla Excel ---
        for trans in transactions: # Iteramos sobre cada transacción para añadirla a la hoja.
            sheet.append([
                trans['idtransaccion'],
                trans['tipo'],
                trans['metododepago'],
                trans['monto_transaccion'],
                trans['fecha_transaccion'],
                trans['estado']
            ])
        try:
            workbook.save(filename) # Guardamos el libro de trabajo de Excel en el archivo especificado.
            messagebox.showinfo("Exportar Excel", f"Reporte de transacciones exportado a {filename}")
        except Exception as e:
            messagebox.showerror("Error de Exportación", f"No se pudo exportar a Excel: {e}")